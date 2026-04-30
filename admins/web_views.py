import openpyxl
from io import BytesIO
from datetime import timedelta
from django.http import JsonResponse
from django.core.cache import cache
from django.conf import settings

from django.db.models import Exists, OuterRef, Sum,  Q

from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib.auth.hashers import check_password, make_password
from django.template.loader import get_template
from django.utils import timezone
from xhtml2pdf import pisa
from members.models import InfoPost
from django.shortcuts import get_object_or_404
from admins.models import AdminUser
from members.models import Member, MemberTransaction, WithdrawalRequest
from admins.utils import (
    log_activity,
    generate_temporary_password,
    generate_temporary_pin,
)
from admins.decorators import (
    admin_session_required,
    super_admin_required,
    member_session_required,
)


from members.services import (
    get_total_contributions,
    get_available_balance,
    get_recent_transactions,
    create_withdrawal_request,
    approve_withdrawal_request,
    reject_withdrawal_request,
)


from members.permissions import forbid_if_no_member_access
from members.constants import (
    DEFAULT_MONTHLY_CONTRIBUTION,
    DEFAULT_PAYMENT_MONTHS,
    MEMBER_STATUS_ACTIVE,
    MEMBER_STATUS_SUSPENDED,
    ADMIN_STATUS_ACTIVE,
    ADMIN_STATUS_SUSPENDED,
    MAX_MEMBER_PIN_ATTEMPTS,
)
from members.payment_services import start_fedapay_payment
from members.services import create_member_record, validate_uploaded_image


def home(request):
    if request.session.get('admin_id'):
        return redirect('/admins/dashboard/')

    if request.session.get('member_id'):
        return redirect('/admins/member-space/')

    return redirect('/admins/login/')


def login(request):
    if request.method == 'POST':


        if is_rate_limited(
            request,
            key_prefix='admin_login',
            max_attempts=settings.LOGIN_RATE_LIMIT_MAX_ATTEMPTS,
            window_seconds=settings.LOGIN_RATE_LIMIT_WINDOW_SECONDS
        ):
            return HttpResponse("Trop de tentatives. Veuillez réessayer dans 10 minutes.")

        email = (request.POST.get('email') or '').strip().lower()
        password = request.POST.get('password') or ''

        try:
            user = AdminUser.objects.get(email=email)

            if user.is_locked:
                return HttpResponse("Compte admin bloqué ❌")

            if user.status == ADMIN_STATUS_SUSPENDED:
                return HttpResponse("Compte suspendu ❌")

            if user.status != ADMIN_STATUS_ACTIVE:
                return HttpResponse("Compte inactif ❌")

            if not check_password(password, user.password):
                user.failed_login_attempts += 1

                if user.failed_login_attempts >= 5:
                    user.is_locked = True
                    user.save(update_fields=['failed_login_attempts', 'is_locked', 'updated_at'])
                    return HttpResponse("Compte admin bloqué après plusieurs tentatives ❌")

                user.save(update_fields=['failed_login_attempts', 'updated_at'])
                remaining = 5 - user.failed_login_attempts
                return HttpResponse(f"Mot de passe incorrect ❌ Il reste {remaining} tentative(s).")

            user.failed_login_attempts = 0
            user.save(update_fields=['failed_login_attempts', 'updated_at'])

            request.session['admin_id'] = user.id
            request.session['admin_role'] = user.role
            request.session['admin_email'] = user.email

            log_activity(
                admin_user=user,
                action='login',
                target_type='admin',
                target_id=user.id,
                details=f"{user.email} s’est connecté"
            )

            if user.must_change_password:
                return redirect('/admins/change-password/')

            return redirect('/admins/dashboard/')

        except AdminUser.DoesNotExist:
            return HttpResponse("Utilisateur introuvable ❌")

    return render(request, 'admins/login.html')




@admin_session_required
def logout(request):
    request.session.flush()
    return redirect('/admins/login/')


@admin_session_required
def change_admin_password(request):
    try:
        admin = AdminUser.objects.get(id=request.session.get('admin_id'))
    except AdminUser.DoesNotExist:
        request.session.flush()
        return redirect('/admins/login/')

    if request.method == 'POST':
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')

        if not new_password or len(new_password) < 6:
            return HttpResponse("Le nouveau mot de passe doit contenir au moins 6 caractères ❌")

        if new_password != confirm_password:
            return HttpResponse("Les deux mots de passe ne correspondent pas ❌")

        admin.password = make_password(new_password)
        admin.must_change_password = False
        admin.save(update_fields=['password', 'must_change_password', 'updated_at'])

        log_activity(
            admin_user=admin,
            action='change_admin_password',
            target_type='admin',
            target_id=admin.id,
            details=f"{admin.email} a changé son mot de passe"
        )

        return redirect('/admins/dashboard/')

    return render(request, 'admins/admin_change_password.html', {
        'admin': admin
    })


@admin_session_required
def dashboard(request):
    admin_id = request.session.get('admin_id')
    admin_email = request.session.get('admin_email')
    role = request.session.get('admin_role')

    total_members = Member.objects.count()
    total_admins = AdminUser.objects.count()

    today = timezone.localdate()
    start_of_week = today - timedelta(days=today.weekday())
    start_of_month = today.replace(day=1)
    start_of_year = today.replace(month=1, day=1)

    global_today = Member.objects.filter(created_at__date=today).count()
    global_this_week = Member.objects.filter(created_at__date__gte=start_of_week).count()
    global_this_month = Member.objects.filter(created_at__date__gte=start_of_month).count()
    global_this_year = Member.objects.filter(created_at__date__gte=start_of_year).count()

    admin_today = Member.objects.filter(created_by_id=admin_id, created_at__date=today).count()
    admin_this_week = Member.objects.filter(created_by_id=admin_id, created_at__date__gte=start_of_week).count()
    admin_this_month = Member.objects.filter(created_by_id=admin_id, created_at__date__gte=start_of_month).count()
    admin_this_year = Member.objects.filter(created_by_id=admin_id, created_at__date__gte=start_of_year).count()

    return render(request, 'admins/dashboard.html', {
        'email': admin_email,
        'role': role,
        'total_members': total_members,
        'total_admins': total_admins,
        'global_today': global_today,
        'global_this_week': global_this_week,
        'global_this_month': global_this_month,
        'global_this_year': global_this_year,
        'admin_today': admin_today,
        'admin_this_week': admin_this_week,
        'admin_this_month': admin_this_month,
        'admin_this_year': admin_this_year,
    })


@admin_session_required
def members_hub(request):
    return render(request, 'admins/members_hub.html')


@admin_session_required
def admins_hub(request):
    return render(request, 'admins/admins_hub.html')


@admin_session_required
def create_info_post(request):
    if request.method == 'POST':
        title = (request.POST.get('title') or '').strip()
        content = (request.POST.get('content') or '').strip()
        video_url = (request.POST.get('video_url') or '').strip() or None
        image = request.FILES.get('image')

        if not title or not content:
            return HttpResponse("Titre et contenu obligatoires ❌")

        admin = AdminUser.objects.get(id=request.session.get('admin_id'))

        InfoPost.objects.create(
            title=title,
            content=content,
            video_url=video_url,
            image=image,
            is_published=True,
            published_at=timezone.now(),
            created_by=admin
        )

        return redirect('/admins/info-posts/')

    return render(request, 'admins/create_info_post.html')


@admin_session_required
def info_post_list(request):
    posts = InfoPost.objects.all().order_by('-created_at')

    return render(request, 'admins/info_post_list.html', {
        'posts': posts
    })



@admin_session_required
def edit_info_post(request, post_id):
    try:
        post = InfoPost.objects.get(id=post_id)
    except InfoPost.DoesNotExist:
        return HttpResponse("Post introuvable ❌")

    if request.method == 'POST':
        title = (request.POST.get('title') or '').strip()
        content = (request.POST.get('content') or '').strip()
        video_url = (request.POST.get('video_url') or '').strip() or None
        image = request.FILES.get('image')

        if not title or not content:
            return HttpResponse("Titre et contenu obligatoires ❌")

        post.title = title
        post.content = content
        post.video_url = video_url

        if image:
            post.image = image

        if not post.published_at:
            post.published_at = timezone.now()

        post.save()

        return redirect('/admins/info-posts/')

    return render(request, 'admins/edit_info_post.html', {
        'post': post
    })


@admin_session_required
def delete_info_post(request, post_id):
    if request.method != 'POST':
        return redirect('/admins/info-posts/')

    try:
        post = InfoPost.objects.get(id=post_id)
    except InfoPost.DoesNotExist:
        return HttpResponse("Post introuvable ❌")

    post.delete()
    return redirect('/admins/info-posts/')    


def member_info_posts(request):
    member_id = request.session.get('member_id')
    if not member_id:
        return redirect('/admins/member-login/')

    try:
        member = Member.objects.get(id=member_id)
    except Member.DoesNotExist:
        request.session.flush()
        return redirect('/admins/member-login/')

    posts = InfoPost.objects.filter(is_published=True).order_by('-published_at', '-created_at')

    return render(request, 'admins/member_info_posts.html', {
        'member': member,
        'posts': posts
    })



@super_admin_required
def admin_list(request):
    search_nim = (request.GET.get('nim') or '').strip()

    admins = AdminUser.objects.all().order_by('-created_at')

    if search_nim:
        admins = admins.filter(nim__icontains=search_nim)

    return render(request, 'admins/admins_list.html', {
        'admins': admins,
        'search_nim': search_nim,
    })


@super_admin_required
def create_admin(request):
    if request.method == 'POST':
        nim = (request.POST.get('nim') or '').strip()
        email = (request.POST.get('email') or '').strip().lower()
        role = (request.POST.get('role') or '').strip()

        if not nim:
            return HttpResponse("Le NIM est obligatoire ❌")

        if role not in ['super_admin', 'admin']:
            return HttpResponse("Rôle invalide ❌")

        try:
            member = Member.objects.get(nim=nim)
        except Member.DoesNotExist:
            return HttpResponse("Aucun membre trouvé avec ce NIM ❌")

        if AdminUser.objects.filter(nim=nim).exists():
            return HttpResponse("Un administrateur existe déjà avec ce NIM ❌")

        if AdminUser.objects.filter(email=email).exists():
            return HttpResponse("Cet email existe déjà ❌")

        if role == 'super_admin' and request.session.get('admin_email') != 'admin@fondaction.com':
            return HttpResponse("Seul le super admin principal peut créer un super admin ❌")

        temporary_password = generate_temporary_password()

        new_admin = AdminUser.objects.create(
            first_name=member.first_name,
            last_name=member.last_name,
            phone=member.phone,
            nim=member.nim,
            email=email,
            password=make_password(temporary_password),
            role=role,
            status=ADMIN_STATUS_ACTIVE,
            must_change_password=True
        )

        current_admin = AdminUser.objects.get(id=request.session.get('admin_id'))

        log_activity(
            admin_user=current_admin,
            action='create_admin',
            target_type='admin',
            target_id=new_admin.id,
            details=f"{current_admin.email} a créé l’admin {new_admin.email} lié au NIM {new_admin.nim}"
        )

        return HttpResponse(
            f"Administrateur créé avec succès ✅<br>"
            f"Mot de passe provisoire : <strong>{temporary_password}</strong><br>"
            f"L’administrateur devra changer ce mot de passe à sa première connexion."
        )

    return render(request, 'admins/create_admin.html')


def get_admin_members_with_payment_status(admin, start_date='', end_date='', payment_status='all'):
    members = Member.objects.filter(created_by=admin).order_by('-created_at')

    if start_date:
        members = members.filter(created_at__date__gte=start_date)

    if end_date:
        members = members.filter(created_at__date__lte=end_date)

    paid_transactions = MemberTransaction.objects.filter(
        member=OuterRef('pk'),
        transaction_type='payment',
        status='success'
    )

    members = members.annotate(
        has_paid=Exists(paid_transactions)
    )

    if payment_status == 'paid':
        members = members.filter(has_paid=True)
    elif payment_status == 'unpaid':
        members = members.filter(has_paid=False)

    return members


def get_global_members_with_payment_status(start_date='', end_date='', payment_status='all'):
    members = Member.objects.select_related('created_by').order_by('-created_at')

    if start_date:
        members = members.filter(created_at__date__gte=start_date)

    if end_date:
        members = members.filter(created_at__date__lte=end_date)

    paid_transactions = MemberTransaction.objects.filter(
        member=OuterRef('pk'),
        transaction_type='payment',
        status='success'
    )

    members = members.annotate(
        has_paid=Exists(paid_transactions),
        total_paid_amount=Sum(
            'transactions__amount',
            filter=Q(
                transactions__transaction_type='payment',
                transactions__status='success'
            )
        )
    )

    if payment_status == 'paid':
        members = members.filter(has_paid=True)
    elif payment_status == 'unpaid':
        members = members.filter(has_paid=False)

    return members


@admin_session_required
def admin_detail(request, admin_id):
    if request.session.get('admin_role') != 'super_admin':
        return HttpResponse("Accès refusé ❌")

    try:
        admin = AdminUser.objects.get(id=admin_id)
    except AdminUser.DoesNotExist:
        return HttpResponse("Admin introuvable ❌")

    start_date = (request.GET.get('start_date') or '').strip()
    end_date = (request.GET.get('end_date') or '').strip()
    payment_status = (request.GET.get('payment_status') or 'all').strip()

    members = get_admin_members_with_payment_status(
        admin=admin,
        start_date=start_date,
        end_date=end_date,
        payment_status=payment_status,
    )

    total_created_members = get_admin_members_with_payment_status(
        admin=admin,
        start_date=start_date,
        end_date=end_date,
        payment_status='all',
    ).count()

    total_paid_members = get_admin_members_with_payment_status(
        admin=admin,
        start_date=start_date,
        end_date=end_date,
        payment_status='paid',
    ).count()

    total_unpaid_members = get_admin_members_with_payment_status(
        admin=admin,
        start_date=start_date,
        end_date=end_date,
        payment_status='unpaid',
    ).count()

    return render(request, 'admins/admin_detail.html', {
        'admin_obj': admin,
        'members': members,
        'total_created_members': total_created_members,
        'total_paid_members': total_paid_members,
        'total_unpaid_members': total_unpaid_members,
        'start_date': start_date,
        'end_date': end_date,
        'payment_status': payment_status,
    })


@super_admin_required
def global_admin_performance(request):
    start_date = (request.GET.get('start_date') or '').strip()
    end_date = (request.GET.get('end_date') or '').strip()
    payment_status = (request.GET.get('payment_status') or 'all').strip()

    members = get_global_members_with_payment_status(
        start_date=start_date,
        end_date=end_date,
        payment_status=payment_status,
    )

    total_created_members = get_global_members_with_payment_status(
        start_date=start_date,
        end_date=end_date,
        payment_status='all',
    ).count()

    total_paid_members = get_global_members_with_payment_status(
        start_date=start_date,
        end_date=end_date,
        payment_status='paid',
    ).count()

    total_unpaid_members = get_global_members_with_payment_status(
        start_date=start_date,
        end_date=end_date,
        payment_status='unpaid',
    ).count()

    total_admins = AdminUser.objects.count()
    total_members_all_time = Member.objects.count()

    total_payments_amount = MemberTransaction.objects.filter(
        member__in=get_global_members_with_payment_status(
            start_date=start_date,
            end_date=end_date,
            payment_status='all',
        ),
        transaction_type='payment',
        status='success'
    ).aggregate(total=Sum('amount'))['total'] or 0

    

    # Version plus robuste pour le résumé par admin
    admins_summary_data = []
    admins = AdminUser.objects.all().order_by('-created_at')

    for admin in admins:
        admin_members = Member.objects.filter(created_by=admin)

        if start_date:
            admin_members = admin_members.filter(created_at__date__gte=start_date)

        if end_date:
            admin_members = admin_members.filter(created_at__date__lte=end_date)

        created_count = admin_members.count()

        paid_count = admin_members.filter(
            transactions__transaction_type='payment',
            transactions__status='success'
        ).distinct().count()

        unpaid_count = created_count - paid_count

        admins_summary_data.append({
            'admin': admin,
            'created_count': created_count,
            'paid_count': paid_count,
            'unpaid_count': unpaid_count,
        })

    admins_summary_data.sort(key=lambda x: x['created_count'], reverse=True)

    return render(request, 'admins/global_admin_performance.html', {
        'members': members,
        'total_created_members': total_created_members,
        'total_paid_members': total_paid_members,
        'total_unpaid_members': total_unpaid_members,
        'total_admins': total_admins,
        'total_members_all_time': total_members_all_time,
        'total_payments_amount': total_payments_amount,
        'admins_summary_data': admins_summary_data,
        'start_date': start_date,
        'end_date': end_date,
        'payment_status': payment_status,
    })


@super_admin_required
def export_global_admin_performance_excel(request):
    start_date = (request.GET.get('start_date') or '').strip()
    end_date = (request.GET.get('end_date') or '').strip()
    payment_status = (request.GET.get('payment_status') or 'all').strip()

    members = get_global_members_with_payment_status(
        start_date=start_date,
        end_date=end_date,
        payment_status=payment_status,
    )

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Performance Globale"

    headers = [
        "Admin créateur",
        "Email admin",
        "Rôle admin",
        "Membre NIM",
        "Nom membre",
        "Prénom membre",
        "Téléphone membre",
        "Ville",
        "Date de création",
        "A payé ?",
        "Montant total payé",
    ]
    sheet.append(headers)

    for member in members:
        admin = member.created_by

        total_paid_amount = MemberTransaction.objects.filter(
            member=member,
            transaction_type='payment',
            status='success'
        ).aggregate(total=Sum('amount'))['total'] or 0

        sheet.append([
            f"{admin.last_name} {admin.first_name}" if admin else "",
            admin.email if admin else "",
            admin.role if admin else "",
            member.nim,
            member.last_name,
            member.first_name,
            member.phone,
            member.city,
            member.created_at.strftime("%d/%m/%Y %H:%M"),
            "Oui" if member.has_paid else "Non",
            float(total_paid_amount),
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename=performance_globale_{payment_status}.xlsx'
    )

    workbook.save(response)
    return response


@super_admin_required
def export_admins_summary_excel(request):
    start_date = (request.GET.get('start_date') or '').strip()
    end_date = (request.GET.get('end_date') or '').strip()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Résumé Admins"

    headers = [
        "NIM Admin",
        "Nom",
        "Prénom",
        "Email",
        "Rôle",
        "Membres créés",
        "Ont payé",
        "N’ont pas payé",
    ]
    sheet.append(headers)

    admins = AdminUser.objects.all().order_by('-created_at')

    for admin in admins:
        admin_members = Member.objects.filter(created_by=admin)

        if start_date:
            admin_members = admin_members.filter(created_at__date__gte=start_date)

        if end_date:
            admin_members = admin_members.filter(created_at__date__lte=end_date)

        created_count = admin_members.count()

        paid_count = admin_members.filter(
            transactions__transaction_type='payment',
            transactions__status='success'
        ).distinct().count()

        unpaid_count = created_count - paid_count

        sheet.append([
            admin.nim,
            admin.last_name,
            admin.first_name,
            admin.email,
            admin.role,
            created_count,
            paid_count,
            unpaid_count,
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=resume_admins.xlsx'

    workbook.save(response)
    return response


@super_admin_required
def reset_admin_password(request, admin_id):
    try:
        admin = AdminUser.objects.get(id=admin_id)
    except AdminUser.DoesNotExist:
        return HttpResponse("Admin introuvable ❌")

    if admin.role == 'super_admin' and request.session.get('admin_email') != 'admin@fondaction.com':
        return HttpResponse("Seul le super admin principal peut réinitialiser le mot de passe d’un super admin ❌")

    temporary_password = generate_temporary_password()
    admin.password = make_password(temporary_password)
    admin.must_change_password = True
    admin.save(update_fields=['password', 'must_change_password', 'updated_at'])

    current_admin = AdminUser.objects.get(id=request.session.get('admin_id'))

    log_activity(
        admin_user=current_admin,
        action='reset_admin_password',
        target_type='admin',
        target_id=admin.id,
        details=f"{current_admin.email} a réinitialisé le mot de passe de {admin.email}"
    )

    return HttpResponse(
        f"Mot de passe réinitialisé avec succès ✅<br>"
        f"Nouveau mot de passe provisoire : <strong>{temporary_password}</strong><br>"
        f"L’administrateur devra le changer à sa prochaine connexion."
    )


@super_admin_required
def suspend_admin(request, admin_id):
    try:
        admin = AdminUser.objects.get(id=admin_id)
    except AdminUser.DoesNotExist:
        return HttpResponse("Admin introuvable ❌")

    if admin.id == request.session.get('admin_id'):
        return HttpResponse("Tu ne peux pas te suspendre toi-même ❌")

    if admin.role == 'super_admin':
        return HttpResponse("Impossible de suspendre un super admin ❌")

    admin.status = ADMIN_STATUS_SUSPENDED
    admin.save(update_fields=['status', 'updated_at'])

    current_admin = AdminUser.objects.get(id=request.session.get('admin_id'))

    log_activity(
        admin_user=current_admin,
        action='suspend_admin',
        target_type='admin',
        target_id=admin.id,
        details=f"{current_admin.email} a suspendu {admin.email}"
    )

    return redirect('/admins/list/')


@super_admin_required
def reactivate_admin(request, admin_id):
    try:
        admin = AdminUser.objects.get(id=admin_id)
    except AdminUser.DoesNotExist:
        return HttpResponse("Admin introuvable ❌")

    admin.status = ADMIN_STATUS_ACTIVE
    admin.save(update_fields=['status', 'updated_at'])

    current_admin = AdminUser.objects.get(id=request.session.get('admin_id'))

    log_activity(
        admin_user=current_admin,
        action='reactivate_admin',
        target_type='admin',
        target_id=admin.id,
        details=f"{current_admin.email} a réactivé {admin.email}"
    )

    return redirect('/admins/list/')


@super_admin_required
def activity_logs(request):
    from logs.models import ActivityLog

    logs = ActivityLog.objects.all().order_by('-created_at')
    return render(request, 'admins/activity_logs.html', {
        'logs': logs
    })


@admin_session_required
def create_member(request):
    if request.method == 'POST':
        try:
            current_admin = AdminUser.objects.get(id=request.session.get('admin_id'))
        except AdminUser.DoesNotExist:
            request.session.flush()
            return redirect('/admins/login/')

        try:
            validate_uploaded_image(request.FILES.get('photo'), ['.jpg', '.jpeg', '.png'], max_size_mb=5)
            validate_uploaded_image(request.FILES.get('id_card_front'), ['.jpg', '.jpeg', '.png'], max_size_mb=5)
            validate_uploaded_image(request.FILES.get('id_card_back'), ['.jpg', '.jpeg', '.png'], max_size_mb=5)
            validate_uploaded_image(request.FILES.get('signature'), ['.png'], max_size_mb=2)

            member = create_member_record(request.POST, request.FILES, current_admin)
        except ValueError as e:
            return HttpResponse(f"{str(e)} ❌")
        except Exception:
            return HttpResponse("Erreur serveur ❌")

        log_activity(
            admin_user=current_admin,
            action='create_member',
            target_type='member',
            target_id=member.id,
            details=f"{current_admin.email} a créé le membre {member.nim}"
        )

        return redirect('/admins/members/')

    return render(request, 'admins/create_member.html')


@admin_session_required
def member_list(request):
    try:
        current_admin = AdminUser.objects.get(id=request.session.get('admin_id'))
    except AdminUser.DoesNotExist:
        request.session.flush()
        return redirect('/admins/login/')

    nim = (request.GET.get('nim') or '').strip()
    phone = (request.GET.get('phone') or '').strip()

    members = Member.objects.all().order_by('-created_at')

    if nim:
        members = members.filter(nim__icontains=nim)

    if phone:
        members = members.filter(phone__icontains=phone)

    return render(request, 'admins/member_list.html', {
        'members': members,
        'admin': current_admin,
        'nim': nim,
        'phone': phone,
    })

@admin_session_required
def member_detail(request, member_id):
    try:
        member = Member.objects.get(id=member_id)
    except Member.DoesNotExist:
        return HttpResponse("Membre introuvable ❌")

    forbidden = forbid_if_no_member_access(request, member)
    if forbidden:
        return forbidden

    transactions = MemberTransaction.objects.filter(
        member=member
    ).order_by('-created_at')

    return render(request, 'admins/member_detail.html', {
        'member': member,
        'transactions': transactions,
    })


def member_login(request):
    if request.method == 'POST':
        nim = (request.POST.get('nim') or '').strip()
        pin = (request.POST.get('pin') or '').strip()

        try:
            member = Member.objects.get(nim=nim)

            if member.status == MEMBER_STATUS_SUSPENDED:
                return HttpResponse("Compte suspendu ❌")

            if member.status != MEMBER_STATUS_ACTIVE:
                return HttpResponse("Compte membre inactif ❌")

            if member.is_locked:
                return HttpResponse("Compte bloqué ❌ Veuillez contacter l’administration pour une réinitialisation.")

            if check_password(pin, member.member_pin):
                member.failed_pin_attempts = 0
                member.save(update_fields=['failed_pin_attempts'])

                request.session['member_id'] = member.id
                request.session['member_nim'] = member.nim
                request.session['member_name'] = f"{member.first_name} {member.last_name}"

                if member.must_change_pin:
                    return redirect('/admins/member-change-pin/')

                return redirect('/admins/member-space/')

            member.failed_pin_attempts += 1

            if member.failed_pin_attempts >= MAX_MEMBER_PIN_ATTEMPTS:
                member.is_locked = True
                member.save(update_fields=['failed_pin_attempts', 'is_locked'])
                return HttpResponse(
                    "Compte bloqué ❌ Vous avez dépassé 3 tentatives incorrectes. "
                    "Veuillez contacter l’administration pour une réinitialisation."
                )

            member.save(update_fields=['failed_pin_attempts'])
            remaining = MAX_MEMBER_PIN_ATTEMPTS - member.failed_pin_attempts
            return HttpResponse(f"PIN incorrect ❌ Il vous reste {remaining} tentative(s).")

        except Member.DoesNotExist:
            return HttpResponse("Membre introuvable ❌")

    return render(request, 'admins/member_login.html')


@member_session_required
def member_change_pin(request):
    try:
        member = Member.objects.get(id=request.session.get('member_id'))
    except Member.DoesNotExist:
        request.session.flush()
        return redirect('/admins/member-login/')

    if request.method == 'POST':
        new_pin = request.POST.get('new_pin')
        confirm_pin = request.POST.get('confirm_pin')

        if not new_pin or not new_pin.isdigit() or len(new_pin) != 5:
            return HttpResponse("Le nouveau PIN doit contenir exactement 5 chiffres ❌")

        if new_pin != confirm_pin:
            return HttpResponse("Les deux PIN ne correspondent pas ❌")

        member.member_pin = make_password(new_pin)
        member.must_change_pin = False
        member.failed_pin_attempts = 0
        member.is_locked = False
        member.save(update_fields=['member_pin', 'must_change_pin', 'failed_pin_attempts', 'is_locked'])

        return redirect('/admins/member-space/')

    return render(request, 'admins/member_change_pin.html')

@member_session_required
def member_space(request):
    member_id = request.session.get('member_id')

    if not member_id:
        return redirect('/admins/member-login/')

    try:
        member = Member.objects.get(id=member_id)
    except Member.DoesNotExist:
        request.session.flush()
        return redirect('/admins/member-login/')

    total_contributions = get_total_contributions(member)
    available_balance = get_available_balance(member)
    recent_transactions = get_recent_transactions(member)

    return render(request, 'admins/member_space.html', {
        'member': member,
        'total_contributions': total_contributions,
        'available_balance': available_balance,
        'recent_transactions': recent_transactions,
    })

@member_session_required
def member_withdrawal(request):
    member_id = request.session.get('member_id')

    if not member_id:
        return redirect('/admins/member-login/')

    try:
        member = Member.objects.get(id=member_id)
    except Member.DoesNotExist:
        request.session.flush()
        return redirect('/admins/member-login/')

    error_message = None
    success_message = None

    if request.method == 'POST':
        amount = (request.POST.get('amount') or '').strip()
        receiver_phone = (request.POST.get('receiver_phone') or '').strip()
        reason = (request.POST.get('reason') or '').strip()
        pin = (request.POST.get('pin') or '').strip()

        try:
            create_withdrawal_request(
                member=member,
                amount=amount,
                receiver_phone=receiver_phone,
                reason=reason,
                pin=pin,
            )
            success_message = "Votre demande de retrait a été envoyée avec succès. Elle est en attente de validation."
        except ValueError as e:
            error_message = str(e)
        except Exception:
            error_message = "Une erreur est survenue lors de la demande de retrait."

    available_balance = get_available_balance(member)

    return render(request, 'admins/member_withdrawal.html', {
        'member': member,
        'available_balance': available_balance,
        'error_message': error_message,
        'success_message': success_message,
    })

@admin_session_required
def withdrawal_requests(request):
    admin_id = request.session.get('admin_id')
    if not admin_id:
        return redirect('/admins/login/')

    requests_list = WithdrawalRequest.objects.select_related(
        'member',
        'transaction',
        'processed_by',
    ).order_by('-created_at')

    return render(request, 'admins/withdrawal_requests.html', {
        'requests_list': requests_list,
    })


@admin_session_required
def approve_withdrawal(request, withdrawal_id):
    admin_id = request.session.get('admin_id')
    if not admin_id:
        return redirect('/admins/login/')

    if request.method != 'POST':
        return redirect('/admins/withdrawal-requests/')

    try:
        admin_user = AdminUser.objects.get(id=admin_id)
    except AdminUser.DoesNotExist:
        request.session.flush()
        return redirect('/admins/login/')

    withdrawal_request = get_object_or_404(WithdrawalRequest, id=withdrawal_id)

    admin_note = (request.POST.get('admin_note') or '').strip()

    try:
        approve_withdrawal_request(
            withdrawal_request=withdrawal_request,
            admin_user=admin_user,
            admin_note=admin_note,
        )
    except ValueError as e:
        return HttpResponse(str(e))
    except Exception:
        return HttpResponse("Erreur lors de la validation du retrait.")

    return redirect('/admins/withdrawal-requests/')


@admin_session_required
def reject_withdrawal(request, withdrawal_id):
    admin_id = request.session.get('admin_id')
    if not admin_id:
        return redirect('/admins/login/')

    if request.method != 'POST':
        return redirect('/admins/withdrawal-requests/')

    try:
        admin_user = AdminUser.objects.get(id=admin_id)
    except AdminUser.DoesNotExist:
        request.session.flush()
        return redirect('/admins/login/')

    withdrawal_request = get_object_or_404(WithdrawalRequest, id=withdrawal_id)

    admin_note = (request.POST.get('admin_note') or '').strip()

    try:
        reject_withdrawal_request(
            withdrawal_request=withdrawal_request,
            admin_user=admin_user,
            admin_note=admin_note,
        )
    except ValueError as e:
        return HttpResponse(str(e))
    except Exception:
        return HttpResponse("Erreur lors du refus du retrait.")

    return redirect('/admins/withdrawal-requests/')





@member_session_required
def member_logout(request):
    request.session.pop('member_id', None)
    request.session.pop('member_nim', None)
    request.session.pop('member_name', None)
    return redirect('/admins/member-login/')


@super_admin_required
def reset_member_pin(request, member_id):
    try:
        member = Member.objects.get(id=member_id)
    except Member.DoesNotExist:
        return HttpResponse("Membre introuvable ❌")

    temporary_pin = generate_temporary_pin()
    member.member_pin = make_password(temporary_pin)
    member.must_change_pin = True
    member.failed_pin_attempts = 0
    member.is_locked = False
    member.save(update_fields=['member_pin', 'must_change_pin', 'failed_pin_attempts', 'is_locked'])

    current_admin = AdminUser.objects.get(id=request.session.get('admin_id'))

    log_activity(
        admin_user=current_admin,
        action='reset_member_pin',
        target_type='member',
        target_id=member.id,
        details=f"{current_admin.email} a réinitialisé le PIN du membre {member.nim}"
    )

    return HttpResponse(
        f"PIN réinitialisé avec succès ✅<br>"
        f"Nouveau PIN provisoire : <strong>{temporary_pin}</strong><br>"
        f"Le membre devra le changer à la prochaine connexion."
    )


@member_session_required
def member_card(request):
    try:
        member = Member.objects.get(id=request.session.get('member_id'))
    except Member.DoesNotExist:
        request.session.flush()
        return redirect('/admins/member-login/')

    return render(request, 'admins/member_card.html', {
        'member': member
    })


@member_session_required
def download_member_card_pdf(request):
    try:
        member = Member.objects.get(id=request.session.get('member_id'))
    except Member.DoesNotExist:
        request.session.flush()
        return redirect('/admins/member-login/')

    template = get_template('admins/member_card_pdf.html')
    html = template.render({
        'member': member,
        'photo_url': request.build_absolute_uri(member.photo.url) if member.photo else '',
        'signature_url': request.build_absolute_uri(member.signature.url) if member.signature else '',
        'logo_url': request.build_absolute_uri('/static/logo-fondaction.png'),
        'card_bg_url': request.build_absolute_uri('/static/card-bg.png'),
        'card_back_bg_url': request.build_absolute_uri('/static/card-back-bg.png'),
    })

    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)

    if pdf.err:
        return HttpResponse("Erreur lors de la génération du PDF ❌")

    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="carte_{member.nim}.pdf"'
    return response

@admin_session_required
def edit_member(request, member_id):
    try:
        member = Member.objects.get(id=member_id)
    except Member.DoesNotExist:
        return HttpResponse("Membre introuvable ❌")

    forbidden = forbid_if_no_member_access(request, member)
    if forbidden:
        return forbidden

    if request.method == 'POST':
        member.first_name = (request.POST.get('first_name') or '').strip()
        member.last_name = (request.POST.get('last_name') or '').strip()
        member.birth_date = request.POST.get('birth_date')
        member.birth_place = (request.POST.get('birth_place') or '').strip()
        member.department = (request.POST.get('department') or '').strip()
        member.commune = (request.POST.get('commune') or '').strip()
        member.city = (request.POST.get('city') or '').strip()
        member.district = (request.POST.get('district') or '').strip()
        member.phone = (request.POST.get('phone') or '').strip() or None
        member.id_card_type = (request.POST.get('id_card_type') or '').strip()
        member.id_card_number = (request.POST.get('id_card_number') or '').strip()
        member.emergency_last_name = (request.POST.get('emergency_last_name') or '').strip() or None
        member.emergency_first_name = (request.POST.get('emergency_first_name') or '').strip() or None
        member.emergency_phone = (request.POST.get('emergency_phone') or '').strip() or None

        if not member.id_card_number:
            return HttpResponse("Le numéro de pièce est obligatoire ❌")

        if Member.objects.filter(id_card_number=member.id_card_number).exclude(id=member.id).exists():
            return HttpResponse("Un autre membre utilise déjà ce numéro de pièce ❌")

        if member.phone and Member.objects.filter(phone=member.phone).exclude(id=member.id).exists():
            return HttpResponse("Un autre membre utilise déjà ce numéro de téléphone ❌")

        try:
            validate_uploaded_image(request.FILES.get('photo'), ['.jpg', '.jpeg', '.png'], max_size_mb=5)
            validate_uploaded_image(request.FILES.get('id_card_front'), ['.jpg', '.jpeg', '.png'], max_size_mb=5)
            validate_uploaded_image(request.FILES.get('id_card_back'), ['.jpg', '.jpeg', '.png'], max_size_mb=5)
            validate_uploaded_image(request.FILES.get('signature'), ['.png'], max_size_mb=2)
        except ValueError as e:
            return HttpResponse(f"{str(e)} ❌")

        if request.FILES.get('photo'):
            member.photo = request.FILES.get('photo')

        if request.FILES.get('id_card_front'):
            member.id_card_front = request.FILES.get('id_card_front')

        if request.FILES.get('id_card_back'):
            member.id_card_back = request.FILES.get('id_card_back')

        if request.FILES.get('signature'):
            member.signature = request.FILES.get('signature')

        member.save()

        try:
            current_admin = AdminUser.objects.get(id=request.session.get('admin_id'))
        except AdminUser.DoesNotExist:
            request.session.flush()
            return redirect('/admins/login/')

        log_activity(
            admin_user=current_admin,
            action='edit_member',
            target_type='member',
            target_id=member.id,
            details=f"{current_admin.email} a modifié le membre {member.nim}"
        )

        return redirect(f'/admins/members/{member.id}/')

    return render(request, 'admins/edit_member.html', {
        'member': member
    })


@admin_session_required
def suspend_member(request, member_id):
    try:
        member = Member.objects.get(id=member_id)
    except Member.DoesNotExist:
        return HttpResponse("Membre introuvable ❌")

    forbidden = forbid_if_no_member_access(request, member)
    if forbidden:
        return forbidden

    member.status = MEMBER_STATUS_SUSPENDED
    member.save(update_fields=['status'])

    current_admin = AdminUser.objects.get(id=request.session.get('admin_id'))

    log_activity(
        admin_user=current_admin,
        action='suspend_member',
        target_type='member',
        target_id=member.id,
        details=f"{current_admin.email} a suspendu le membre {member.nim}"
    )

    return redirect(f'/admins/members/{member.id}/')


@admin_session_required
def activate_member(request, member_id):
    try:
        member = Member.objects.get(id=member_id)
    except Member.DoesNotExist:
        return HttpResponse("Membre introuvable ❌")

    forbidden = forbid_if_no_member_access(request, member)
    if forbidden:
        return forbidden

    member.status = MEMBER_STATUS_ACTIVE
    member.save(update_fields=['status'])

    current_admin = AdminUser.objects.get(id=request.session.get('admin_id'))

    log_activity(
        admin_user=current_admin,
        action='activate_member',
        target_type='member',
        target_id=member.id,
        details=f"{current_admin.email} a réactivé le membre {member.nim}"
    )

    return redirect(f'/admins/members/{member.id}/')


@admin_session_required
def member_creation_history(request):
    admin_id = request.session.get('admin_id')
    role = request.session.get('admin_role')

    if role == 'super_admin':
        members = Member.objects.all().order_by('-created_at')
    else:
        members = Member.objects.filter(created_by_id=admin_id).order_by('-created_at')

    return render(request, 'admins/member_creation_history.html', {
        'members': members,
        'role': role,
    })


@super_admin_required
def admin_performance_detail(request, admin_id):
    try:
        admin = AdminUser.objects.get(id=admin_id)
    except AdminUser.DoesNotExist:
        return HttpResponse("Admin introuvable ❌")

    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    members = Member.objects.filter(created_by=admin).order_by('-created_at')

    if start_date:
        members = members.filter(created_at__date__gte=start_date)

    if end_date:
        members = members.filter(created_at__date__lte=end_date)

    total_members = members.count()

    return render(request, 'admins/admin_performance_detail.html', {
        'admin_obj': admin,
        'members': members,
        'total_members': total_members,
        'start_date': start_date,
        'end_date': end_date,
    })


@super_admin_required
def export_admin_performance_excel(request, admin_id):
    try:
        admin = AdminUser.objects.get(id=admin_id)
    except AdminUser.DoesNotExist:
        return HttpResponse("Admin introuvable ❌")

    start_date = (request.GET.get('start_date') or '').strip()
    end_date = (request.GET.get('end_date') or '').strip()
    payment_status = (request.GET.get('payment_status') or 'all').strip()

    members = get_admin_members_with_payment_status(
        admin=admin,
        start_date=start_date,
        end_date=end_date,
        payment_status=payment_status,
    )

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Performance Admin"

    headers = [
        "Admin NIM",
        "Admin Nom",
        "Admin Prénom",
        "Admin Email",
        "Rôle",
        "Membre NIM",
        "Membre Nom",
        "Membre Prénom",
        "Téléphone membre",
        "Ville",
        "Date de création",
        "A payé ?",
    ]
    sheet.append(headers)

    for member in members:
        sheet.append([
            admin.nim,
            admin.last_name,
            admin.first_name,
            admin.email,
            admin.role,
            member.nim,
            member.last_name,
            member.first_name,
            member.phone,
            member.city,
            member.created_at.strftime("%d/%m/%Y %H:%M"),
            "Oui" if member.has_paid else "Non",
        ])

    filename_suffix = payment_status
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename=performance_admin_{admin.id}_{filename_suffix}.xlsx'
    )

    workbook.save(response)
    return response


@admin_session_required
def export_members_excel(request):
    role = request.session.get('admin_role')
    admin_id = request.session.get('admin_id')

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Membres"

    headers = [
        "NIM",
        "Nom",
        "Prénom",
        "Téléphone",
        "Ville",
        "Date de création",
        "Créé par"
    ]
    sheet.append(headers)

    members = Member.objects.all().order_by('-created_at')
    if role != 'super_admin':
        members = members.filter(created_by_id=admin_id)

    for member in members:
        sheet.append([
            member.nim,
            member.last_name,
            member.first_name,
            member.phone,
            member.city,
            member.created_at.strftime("%d/%m/%Y %H:%M"),
            member.created_by.email if member.created_by else ""
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=liste_membres.xlsx'

    workbook.save(response)
    return response


@super_admin_required
def export_admins_excel(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Admins"

    headers = [
        "ID",
        "Nom",
        "Prénom",
        "Email",
        "Téléphone",
        "Rôle",
        "Statut",
        "Date de création"
    ]
    sheet.append(headers)

    admins = AdminUser.objects.all().order_by('-created_at')

    for admin in admins:
        sheet.append([
            admin.id,
            admin.last_name,
            admin.first_name,
            admin.email,
            admin.phone,
            admin.role,
            admin.status,
            admin.created_at.strftime("%d/%m/%Y %H:%M") if admin.created_at else ""
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=liste_admins.xlsx'

    workbook.save(response)
    return response


@member_session_required
def member_payment(request):
    try:
        member = Member.objects.get(id=request.session.get('member_id'))
    except Member.DoesNotExist:
        request.session.flush()
        return redirect('/admins/member-login/')

    now = timezone.localtime()
    french_months = {
        1: "Janvier",
        2: "Février",
        3: "Mars",
        4: "Avril",
        5: "Mai",
        6: "Juin",
        7: "Juillet",
        8: "Août",
        9: "Septembre",
        10: "Octobre",
        11: "Novembre",
        12: "Décembre",
    }

    current_month_label = f"{french_months[now.month]} {now.year}"

    return render(request, 'admins/payment_page.html', {
        'member': member,
        'current_month_label': current_month_label,
        'amount': DEFAULT_MONTHLY_CONTRIBUTION,
    })


@member_session_required
def start_member_payment_view(request):
    if request.method != 'POST':
        return redirect('/admins/member-payment/')

    try:
        member = Member.objects.get(id=request.session.get('member_id'))
    except Member.DoesNotExist:
        request.session.flush()
        return redirect('/admins/member-login/')

    try:
        payment_url, _attempt = start_fedapay_payment(
            member=member,
            amount=DEFAULT_MONTHLY_CONTRIBUTION,
            months=DEFAULT_PAYMENT_MONTHS,
            callback_url=request.build_absolute_uri('/admins/payment-return/'),
        )
        return redirect(payment_url)
    except Exception as e:
        return HttpResponse(f"Erreur paiement : {str(e)} ❌")


def start_member_payment(request):
    return start_member_payment_view(request)


@member_session_required
def payment_return(request):
    status = request.GET.get('status', '')
    transaction_id = request.GET.get('id', '')

    return render(request, 'admins/payment_return.html', {
        'status': status,
        'transaction_id': transaction_id,
    })


@member_session_required
def member_transactions(request):
    try:
        member = Member.objects.get(id=request.session.get('member_id'))
    except Member.DoesNotExist:
        request.session.flush()
        return redirect('/admins/member-login/')

    transactions = MemberTransaction.objects.filter(
        member=member
    ).order_by('-created_at')

    return render(request, 'admins/member_transactions.html', {
        'member': member,
        'transactions': transactions
    }) 

@member_session_required
def member_transaction_detail(request, transaction_id):
    member_id = request.session.get('member_id')

    if not member_id:
        return redirect('/admins/member-login/')

    try:
        member = Member.objects.get(id=member_id)
    except Member.DoesNotExist:
        request.session.flush()
        return redirect('/admins/member-login/')

    transaction = get_object_or_404(
        MemberTransaction,
        id=transaction_id,
        member=member
    )

    withdrawal_request = None
    try:
        withdrawal_request = transaction.withdrawal_request
    except WithdrawalRequest.DoesNotExist:
        withdrawal_request = None
    except AttributeError:
        withdrawal_request = None

    return render(request, 'admins/member_transaction_detail.html', {
        'member': member,
        'transaction': transaction,
        'withdrawal_request': withdrawal_request,
    })   

@member_session_required
def member_profile(request):
    try:
        member = Member.objects.get(id=request.session.get('member_id'))
    except Member.DoesNotExist:
        request.session.flush()
        return redirect('/admins/member-login/')

    return render(request, 'admins/member_profile.html', {
        'member': member
    })

@member_session_required
def member_settings(request):
    try:
        member = Member.objects.get(id=request.session.get('member_id'))
    except Member.DoesNotExist:
        request.session.flush()
        return redirect('/admins/member-login/')

    return render(request, 'admins/member_settings.html', {
        'member': member
    })



@admin_session_required
def search_transactions(request):
    admin_id = request.session.get('admin_id')

    if not admin_id:
        return redirect('/admins/login/')

    query = (request.GET.get('q') or '').strip()
    transactions = []

    if query:
        transactions = MemberTransaction.objects.select_related('member').filter(
            receipt_number__icontains=query
        ).order_by('-created_at')

    return render(request, 'admins/search_transactions.html', {
        'query': query,
        'transactions': transactions,
    })

def member_assistance(request):
    return render(request, 'admins/member_assistance.html')


def visitor_home(request):
    return render(request, 'admins/visitor_home.html')

def secure_registration_page(request):
    return render(request, 'admins/secure_registration.html')

def about_page(request):
    return render(request, 'admins/about.html')

def privacy_policy_page(request):
    return render(request, 'admins/privacy_policy.html')


def terms_services_page(request):
    return render(request, 'admins/terms_services.html')

def contact_page(request):
    return render(request, 'admins/contact.html')


def root_router(request):
    host = request.get_host().split(':')[0]

    if host == "admin.fondactionsarl.com":
        return redirect('/admins/login/')

    if host == "api.fondactionsarl.com":
        return JsonResponse({
            "success": True,
            "message": "FondAction API is running"
        })

    return redirect('/admins/visiteur/')

def robots_txt(request):
    host = request.get_host()

    if host.startswith("admin.") or host.startswith("api."):
        return HttpResponse("User-agent: *\nDisallow: /", content_type="text/plain")

    return HttpResponse("User-agent: *\nAllow: /", content_type="text/plain")



def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')

    if x_forwarded_for:
        return x_forwarded_for.split(',')[0].strip()

    return request.META.get('REMOTE_ADDR', '')


def is_rate_limited(request, key_prefix, max_attempts=5, window_seconds=600):
    ip = get_client_ip(request)
    cache_key = f"{key_prefix}:{ip}"

    attempts = cache.get(cache_key, 0)

    if attempts >= max_attempts:
        return True

    cache.set(cache_key, attempts + 1, timeout=window_seconds)
    return False


def sitemap_xml(request):
    xml = """<?xml version="1.0" encoding="UTF-8"?>
<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">
    <url>
        <loc>https://fondactionsarl.com/</loc>
        <priority>1.0</priority>
    </url>
    <url>
        <loc>https://fondactionsarl.com/admins/about/</loc>
        <priority>0.8</priority>
    </url>
    <url>
        <loc>https://fondactionsarl.com/admins/contact/</loc>
        <priority>0.8</priority>
    </url>
    <url>
        <loc>https://fondactionsarl.com/admins/privacy-policy/</loc>
        <priority>0.5</priority>
    </url>
    <url>
        <loc>https://fondactionsarl.com/admins/terms-services/</loc>
        <priority>0.5</priority>
    </url>
</urlset>
"""
    return HttpResponse(xml, content_type="application/xml")